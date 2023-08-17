import openpyxl
import requests
from bs4 import BeautifulSoup


# 透過貨幣取得其匯率
def crawler(currency):
    url = f'https://www.google.com/search?q={currency}'
    user_agent = 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/104.0.5112.102 Safari/537.36 Edg/104.0.1293.70'
    res = requests.get(url, headers={
        'user-agent': user_agent
    })
    soup = BeautifulSoup(res.text, 'html.parser')

    forex_rate = soup.find('span', class_='DFlfde SwHCTb')
    return forex_rate.text if forex_rate else None


workbook = openpyxl.load_workbook('匯率即時更新.xlsx')
sheet = workbook['即時匯率']
max_row = sheet.max_row

currency_list = []
# 讀取目標 excel 現有的所有貨幣，並存入 currency_list。
for r in range(3, max_row+1):
    # 讀取 row = 3～max_row, column = 1 的儲存格值
    currency = sheet.cell(row=r, column=1).value
    currency_list.append(currency)

print(currency_list)  # ['美金', '日幣', '歐元',...]

now_row = 2
# 將 currency_list 中的每個元素（貨幣）都呼叫一次匯率爬蟲
for currency in currency_list:
    result = crawler(currency)
    sheet.cell(row=now_row, column=2).value = result
    now_row += 1

workbook.save('匯率即時更新.xlsx')
