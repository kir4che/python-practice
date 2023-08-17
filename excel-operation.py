import openpyxl

# 讀檔
workbook = openpyxl.load_workbook('test.xlsx')

# 新增工作表
workbook.create_sheet('工作表 2')
# 查看全部工作表
print(workbook.sheetnames)  # ['工作表 1', '工作表 2']
# 選取特定工作表
print(workbook['工作表 1'])  # <Worksheet "工作表 1">

sheet = workbook['工作表 1']
# 最大列數
print(sheet.max_row)  # 23
# 最大行數
print(sheet.max_column)  # 7
# 凍結窗格
sheet.freeze_panes = 'A2'
# 指定列／行的寬高
sheet.column_dimensions['B'].width = 20
# 合併儲存格
sheet.merge_cells('A1:D3')  # merge 12 cells

# 修改該儲存格內容
sheet['A4'].value = 'Hello'  # 可省略 .value

# 取得該儲存格的內容
print(sheet['A4'].value)  # Hello
print(sheet.cell(row=3, column=2).value)  # None

# 取得多個儲存格的內容
sheet['B']
# (<Cell '工作表1'.B1>, <Cell '工作表1'.B2>, <Cell '工作表1'.B3>, <Cell '工作表1'.B4>)
sheet['B2:E4']
# ((<Cell '工作表1'.B2>, <Cell '工作表1'.C2>, <Cell '工作表1'.D2>, <Cell '工作表1'.E2>), (<Cell '工作表1'.B3>, <Cell '工作表1'.C3>, <Cell '工作表1'.D3>, <Cell '工作表1'.E3>), (<Cell '工作表1'.B4>, <Cell '工作表1'.C4>, <Cell '工作表1'.D4>, <Cell '工作表1'.E4>))
list(sheet.rows)
# [(<Cell '工作表1'.A1>, <Cell '工作表1'.B1>, <Cell '工作表1'.C1>, <Cell '工作表1'.D1>, <Cell '工作表1'.E1>), (<Cell '工作表1'.A2>, <Cell '工作表1'.B2>, <Cell '工作表1'.C2>, <Cell '工作表1'.D2>, <Cell '工作表1'.E2>), (<Cell '工作表1'.A3>, <Cell '工作表1'.B3>, <Cell '工作表1'.C3>, <Cell '工作表1'.D3>, <Cell '工作表1'.E3>), (<Cell '工作表1'.A4>, <Cell '工作表1'.B4>, <Cell '工作表1'.C4>, <Cell '工作表1'.D4>, <Cell '工作表1'.E4>)]
list(sheet.iter_rows(1, 6, 1, 5))  # 先列後欄
# [(<Cell '工作表1'.A1>, <Cell '工作表1'.B1>, <Cell '工作表1'.C1>, <Cell '工作表1'.D1>, <Cell '工作表1'.E1>), (<Cell '工作表1'.A2>, <Cell '工作表1'.B2>, <Cell '工作表1'.C2>, <Cell '工作表1'.D2>, <Cell '工作表1'.E2>), (<Cell '工作表1'.A3>, <Cell '工作表1'.B3>, <Cell '工作表1'.C3>, <Cell '工作表1'.D3>, <Cell '工作表1'.E3>), (<Cell '工作表1'.A4>, <Cell '工作表1'.B4>, <Cell '工作表1'.C4>, <Cell '工作表1'.D4>, <Cell '工作表1'.E4>), (<Cell '工作表1'.A5>, <Cell '工作表1'.B5>, <Cell '工作表1'.C5>, <Cell '工作表1'.D5>, <Cell '工作表1'.E5>), (<Cell '工作表1'.A6>, <Cell '工作表1'.B6>, <Cell '工作表1'.C6>, <Cell '工作表1'.D6>, <Cell '工作表1'.E6>)]
